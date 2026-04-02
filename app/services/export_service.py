import csv
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.core.exceptions import ExportError
from app.models.listing_model import Listing
from app.repositories.listings_repository import ListingRepository


def _listing_to_dict(listing: Listing) -> dict:
    return {
        "id": str(listing.id),
        "partner_id": listing.partner_id,
        "source_partner": listing.source_partner,
        "source_url": listing.source_url,
        "title": listing.title,
        "listing_type": listing.listing_type,
        "property_type": listing.property_type,
        "typology": listing.typology,
        "bedrooms": listing.bedrooms,
        "bathrooms": listing.bathrooms,
        "floor": listing.floor,
        "price_amount": float(listing.price_amount) if listing.price_amount is not None else None,
        "price_currency": listing.price_currency,
        "price_per_m2": float(listing.price_per_m2) if listing.price_per_m2 is not None else None,
        "area_useful_m2": listing.area_useful_m2,
        "area_gross_m2": listing.area_gross_m2,
        "area_land_m2": listing.area_land_m2,
        "district": listing.district,
        "county": listing.county,
        "parish": listing.parish,
        "full_address": listing.full_address,
        "latitude": listing.latitude,
        "longitude": listing.longitude,
        "has_garage": listing.has_garage,
        "has_elevator": listing.has_elevator,
        "has_balcony": listing.has_balcony,
        "has_air_conditioning": listing.has_air_conditioning,
        "has_pool": listing.has_pool,
        "energy_certificate": listing.energy_certificate,
        "construction_year": listing.construction_year,
        "advertiser": listing.advertiser,
        "contacts": listing.contacts,
        "description": listing.description,
        "enriched_description": listing.enriched_description,
        "description_quality_score": listing.description_quality_score,
        "meta_description": listing.meta_description,
        "created_at": listing.created_at.isoformat() if listing.created_at else None,
        "updated_at": listing.updated_at.isoformat() if listing.updated_at else None,
    }


class ExportService:

    @staticmethod
    async def _load_rows(db: AsyncSession, filters: dict) -> list[dict]:
        listings = await ListingRepository.get_listings_for_export(
            db, filters, limit=settings.export_max_rows + 1
        )
        if len(listings) > settings.export_max_rows:
            raise ExportError(
                f"Export exceeds maximum row limit of {settings.export_max_rows}. Refine filters and try again."
            )
        return [_listing_to_dict(l) for l in listings]

    @staticmethod
    async def export_csv(db: AsyncSession, filters: dict) -> io.StringIO:
        rows = await ExportService._load_rows(db, filters)
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        output.seek(0)
        return output

    @staticmethod
    async def export_json(db: AsyncSession, filters: dict) -> str:
        rows = await ExportService._load_rows(db, filters)
        return json.dumps(rows, ensure_ascii=False, indent=2)

    @staticmethod
    async def export_excel(db: AsyncSession, filters: dict) -> io.BytesIO:
        rows = await ExportService._load_rows(db, filters)
        wb = Workbook()
        ws = wb.active
        ws.title = "Listings"
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            bold = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold
            for row_dict in rows:
                ws.append(list(row_dict.values()))
            for i, col_cells in enumerate(ws.columns, 1):
                max_len = max(
                    (len(str(cell.value)) for cell in col_cells if cell.value is not None),
                    default=0,
                )
                ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)
        else:
            ws.append(["No data found"])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output