"""Export API router — download listings as CSV, JSON, or Excel.
/api/v1/export
"""

import csv
import io
import json
from datetime import datetime
from decimal import Decimal

from uuid import UUID

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_db
from app.api.responses import ERROR_RESPONSES
from app.api.v1._filters import apply_listing_filters
from app.config import settings
from app.core.exceptions import ExportError
from app.models.listing_model import Listing

router = APIRouter()


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _listing_to_dict(listing: Listing) -> dict:
    """Convert a listing ORM object to a flat dict suitable for export."""
    return {
        "id": str(listing.id),
        "partner_id": listing.partner_id,
        "source_partner": listing.source_partner,
        "source_url": listing.source_url,
        "title": listing.title,
        "listing_type": listing.listing_type,
        "property_type": listing.property_type,
        "typology": listing.typology,
        "bedrooms": listing.bedrooms,
        "bathrooms": listing.bathrooms,
        "floor": listing.floor,
        "price_amount": float(listing.price_amount) if listing.price_amount is not None else None,
        "price_currency": listing.price_currency,
        "price_per_m2": float(listing.price_per_m2) if listing.price_per_m2 is not None else None,
        "area_useful_m2": listing.area_useful_m2,
        "area_gross_m2": listing.area_gross_m2,
        "area_land_m2": listing.area_land_m2,
        "district": listing.district,
        "county": listing.county,
        "parish": listing.parish,
        "full_address": listing.full_address,
        "latitude": listing.latitude,
        "longitude": listing.longitude,
        "has_garage": listing.has_garage,
        "has_elevator": listing.has_elevator,
        "has_balcony": listing.has_balcony,
        "has_air_conditioning": listing.has_air_conditioning,
        "has_pool": listing.has_pool,
        "energy_certificate": listing.energy_certificate,
        "construction_year": listing.construction_year,
        "advertiser": listing.advertiser,
        "contacts": listing.contacts,
        "description": listing.description,
        "enriched_description": listing.enriched_description,
        "description_quality_score": listing.description_quality_score,
        "meta_description": listing.meta_description,
        "created_at": listing.created_at.isoformat() if listing.created_at else None,
        "updated_at": listing.updated_at.isoformat() if listing.updated_at else None,
    }


async def _load_export_rows(db: AsyncSession, filters: dict) -> list[Listing]:
    """Load export rows with a hard cap to protect API memory usage."""
    query = apply_listing_filters(select(Listing), **filters).order_by(Listing.created_at.desc()).limit(settings.export_max_rows + 1)
    listings = (await db.execute(query)).scalars().all()
    if len(listings) > settings.export_max_rows:
        raise ExportError(
            f"Export exceeds maximum row limit of {settings.export_max_rows}. Refine filters and try again."
        )
    return listings


# ---------------------------------------------------------------------------
# Shared query params (DRY — avoids repeating 7 Query() declarations)
# ---------------------------------------------------------------------------

def _export_filters(
    district: str | None = Query(None),
    county: str | None = Query(None),
    parish: str | None = Query(None),
    property_type: str | None = Query(None),
    typology: str | None = Query(None),
    listing_type: str | None = Query(None, pattern="^(sale|rent)$"),
    source_partner: str | None = Query(None),
    scrape_job_id: UUID | None = Query(None),
    price_min: Decimal | None = Query(None),
    price_max: Decimal | None = Query(None),
    area_min: float | None = Query(None),
    area_max: float | None = Query(None),
    bedrooms_min: int | None = Query(None),
    bedrooms_max: int | None = Query(None),
    has_garage: bool | None = Query(None),
    has_pool: bool | None = Query(None),
    has_elevator: bool | None = Query(None),
    created_after: datetime | None = Query(None),
    created_before: datetime | None = Query(None),
    search: str | None = Query(None),
):
    return dict(
        district=district,
        county=county,
        parish=parish,
        property_type=property_type,
        typology=typology,
        listing_type=listing_type,
        source_partner=source_partner,
        scrape_job_id=scrape_job_id,
        price_min=price_min,
        price_max=price_max,
        area_min=area_min,
        area_max=area_max,
        bedrooms_min=bedrooms_min,
        bedrooms_max=bedrooms_max,
        has_garage=has_garage,
        has_pool=has_pool,
        has_elevator=has_elevator,
        created_after=created_after,
        created_before=created_before,
        search=search,
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get(
    "/csv",
    summary="Export CSV",
    operation_id="export_csv",
    responses={
        200: {"content": {"text/csv": {}}, "description": "CSV file download"},
        **ERROR_RESPONSES,
    },
)
async def export_csv(
    db: AsyncSession = Depends(get_db),
    filters: dict = Depends(_export_filters),
):
    """Export filtered listings as CSV."""
    listings = await _load_export_rows(db, filters)
    output = io.StringIO()

    if listings:
        rows = [_listing_to_dict(l) for l in listings]
        writer = csv.DictWriter(output, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=listings_export.csv"},
    )


@router.get(
    "/json",
    summary="Export JSON",
    operation_id="export_json",
    responses={
        200: {"content": {"application/json": {}}, "description": "JSON file download"},
        **ERROR_RESPONSES,
    },
)
async def export_json(
    db: AsyncSession = Depends(get_db),
    filters: dict = Depends(_export_filters),
):
    """Export filtered listings as JSON."""
    listings = await _load_export_rows(db, filters)
    rows = [_listing_to_dict(l) for l in listings]
    content = json.dumps(rows, ensure_ascii=False, indent=2)

    return StreamingResponse(
        iter([content]),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=listings_export.json"},
    )


@router.get(
    "/excel",
    summary="Export Excel",
    operation_id="export_excel",
    responses={
        200: {
            "content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}},
            "description": "Excel (.xlsx) file download",
        },
        **ERROR_RESPONSES,
    },
)
async def export_excel(
    db: AsyncSession = Depends(get_db),
    filters: dict = Depends(_export_filters),
):
    """Export filtered listings as Excel (.xlsx)."""
    listings = await _load_export_rows(db, filters)
    rows = [_listing_to_dict(l) for l in listings]

    wb = Workbook()
    ws = wb.active
    ws.title = "Listings"

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)

        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold

        for row_dict in rows:
            ws.append(list(row_dict.values()))

        # Auto-width columns (capped at 50)
        for i, col_cells in enumerate(ws.columns, 1):
            max_len = max(
                (len(str(cell.value)) for cell in col_cells if cell.value is not None),
                default=0,
            )
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)
    else:
        ws.append(["No data found"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=listings_export.xlsx"},
    )